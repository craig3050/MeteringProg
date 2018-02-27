from openpyxl import Workbook
from openpyxl import load_workbook


# import sqlite3
# conn = sqlite3.connect('import_data.db')


# Part 1 - open the excel file
def load_excel_file(file_path):
    workbook = load_workbook(file_path)
    sheet_names = []
    for i in workbook.get_sheet_names():
        sheet_names.append(i)
    print(sheet_names)
    number = 1  # Starting the list from one as it's easier for people to relate
    for item in sheet_names:
        print(str(number) + " - " + item)
        number += 1
    sheet_select = int(input("Enter the number of the sheet to be imported: "))
    sheet_select -= 1  # take 1 from the value as list addresses start from zero
    worksheet = workbook[sheet_names[sheet_select]]  # open the selected sheet

    #Print all the column headings
    column_data = ""
    column_names = []
    column_number = 1
    while column_data != None:
        column_data = (worksheet.cell(row=1, column=column_number).value)
        column_number += 1


#FIXME - change the print statement so it goes down a column with a gap after the last one.
    column_number -=1 #remove one from the number as last value is always None
    row_number = 5
    for column_iteration in range(column_number):
        column_names = [] #clear this again
        column_iteration += 1
        for row_iteration in range(row_number):
            row_iteration +=1
            column_data = (worksheet.cell(row=row_iteration, column=column_iteration).value)
            column_names.append(column_data)
        for item in column_names:
            print (item)
        print ("\n")
    input("Press Enter to exit")


if __name__ == '__main__':
    print("This programme does stuff with metering data!\n\n")
    file_path = input("Enter the Path of the document to be processed: ")
    print("Printing directory contents: \n")
    load_excel_file(file_path)
